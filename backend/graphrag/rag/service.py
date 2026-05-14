import os
import json
import glob
import shutil
import asyncio
import pathlib
import tempfile
from typing import List, Dict, Optional
from datetime import datetime

from loguru import logger
from sqlalchemy.ext.asyncio import AsyncSession

from graphrag.rag.socket_manager import manager
from graphrag.config import get_config

from graphrag.rag.db_service import (
    KnowledgeBaseService,
    KnowledgeBaseFileService,
    KnowledgeGraphService,
)
from graphrag.rag.model import KnowledgeBaseFile as KBFileModel, KnowledgeGraph as KGraphModel
from graphrag.rag.schema import (
    KnowledgeBaseFileResponse,
    FileUploadResponse,
)

path = pathlib.Path(__file__).parent.parent

try:
    from graphrag.models.constructor import kt_gen as constructor
    from graphrag.models.retriever import (
        agentic_decomposer as decomposer,
        enhanced_kt_retriever as retriever,
    )
    GRAPHRAG_AVAILABLE = True
    logger.info("GraphRAG components loaded successfully")
except ImportError as e:
    GRAPHRAG_AVAILABLE = False
    constructor = None
    decomposer = None
    retriever = None
    logger.error(f"GraphRAG components not available: {e}")

config = None

DEFAULT_SCHEMA = {
    "Nodes": [
        "person", "location", "organization", "event",
        "object", "concept", "time_period", "creative_work",
        "biological_entity", "natural_phenomenon",
    ],
    "Relations": [
        "is_a", "part_of", "located_in", "created_by", "used_by",
        "participates_in", "related_to", "belongs_to", "influences",
        "precedes", "arrives_in", "comparable_to",
    ],
    "Attributes": [
        "name", "date", "size", "type", "description",
        "status", "quantity", "value", "position", "duration", "time",
    ],
}


def get_config_instance():
    global config
    if config is None:
        config = get_config()
    return config


def _ensure_schema(schema_json: Optional[dict]) -> dict:
    return schema_json or DEFAULT_SCHEMA


async def send_progress_update(client_id: str, stage: str, progress: int, message: str):
    await manager.send_message(
        {
            "type": "progress",
            "stage": stage,
            "progress": progress,
            "message": message,
            "timestamp": datetime.now().isoformat(),
        },
        client_id,
    )


def _clear_cache_files_sync(kb_id: str, file_id: str):
    try:
        cache_key = file_id
        faiss_cache_dir = path / f"retriever/faiss_cache_new/{cache_key}"
        if os.path.exists(faiss_cache_dir):
            shutil.rmtree(faiss_cache_dir)
        chunk_file = path / f"output/chunks/{cache_key}.txt"
        if os.path.exists(chunk_file):
            os.remove(chunk_file)
        graph_file = path / f"output/graphs/{cache_key}_new.json"
        if os.path.exists(graph_file):
            os.remove(graph_file)
        for pattern in [
            path / f"output/logs/{cache_key}_*.log",
            path / f"output/chunks/{cache_key}_*",
            path / f"output/graphs/{cache_key}_*",
        ]:
            for fp in glob.glob(str(pattern)):
                try:
                    if os.path.isfile(fp):
                        os.remove(fp)
                    elif os.path.isdir(fp):
                        shutil.rmtree(fp)
                except Exception as e:
                    logger.warning(f"Failed to clear {fp}: {e}")
        logger.info(f"Cache cleanup completed for file: {file_id}")
    except Exception as e:
        logger.error(f"Error clearing cache files for {file_id}: {e}")


async def clear_cache_files(kb_id: str, file_id: str):
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _clear_cache_files_sync, kb_id, file_id)


def _extract_text_from_document_sync(file_path: str, file_ext: str) -> str:
    try:
        if file_ext in [".txt", ".md"]:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        elif file_ext == ".pdf":
            try:
                import PyPDF2
                text = ""
                with open(file_path, "rb") as f:
                    for page in PyPDF2.PdfReader(f).pages:
                        text += page.extract_text() + "\n"
                return text
            except ImportError:
                try:
                    import pdfplumber
                    text = ""
                    with pdfplumber.open(file_path) as pdf:
                        for page in pdf.pages:
                            text += page.extract_text() + "\n"
                    return text
                except ImportError:
                    return f"[PDF content from {os.path.basename(file_path)}]"
        elif file_ext in [".doc", ".docx"]:
            try:
                import docx
                return "\n".join(p.text for p in docx.Document(file_path).paragraphs)
            except ImportError:
                return f"[Word content from {os.path.basename(file_path)}]"
    except Exception as e:
        return f"[Error extracting content from {os.path.basename(file_path)}: {str(e)}]"
    return ""


async def extract_text_from_document(file_path: str, file_ext: str) -> str:
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _extract_text_from_document_sync, file_path, file_ext)


def _extract_text_from_spreadsheet_sync(file_path: str, file_ext: str):
    try:
        if file_ext == ".csv":
            import csv
            text = ""
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                for row in csv.reader(f):
                    text += ", ".join(row) + "\n"
            return text
        elif file_ext in [".xls", ".xlsx"]:
            try:
                import pandas as pd
                return pd.read_excel(file_path).to_string(index=False)
            except ImportError:
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                text = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    text += f"Sheet: {sheet}\n"
                    for row in ws.iter_rows():
                        text += ", ".join(str(c.value) if c.value else "" for c in row) + "\n"
                    text += "\n"
                return text
    except Exception as e:
        return f"[Error: {str(e)}]"
    return ""


async def extract_text_from_spreadsheet(file_path: str, file_ext: str):
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _extract_text_from_spreadsheet_sync, file_path, file_ext)


async def upload_files_to_kb(
    kb_id: str, files: List, schema_json: Optional[dict], db: AsyncSession
) -> List[FileUploadResponse]:
    if not files:
        raise Exception("No files provided")

    effective_schema = _ensure_schema(schema_json)
    results = []

    for file in files:
        content_bytes = await file.read()
        file_ext = os.path.splitext(file.filename)[1].lower()
        file_size = len(content_bytes)

        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
            tmp.write(content_bytes)
            tmp_path = tmp.name

        try:
            if file_ext in [".csv", ".xls", ".xlsx"]:
                text = await extract_text_from_spreadsheet(tmp_path, file_ext)
            else:
                text = await extract_text_from_document(tmp_path, file_ext)
        finally:
            os.unlink(tmp_path)

        from app.base_model import generate_nanoid

        file_record = KBFileModel(
            id=generate_nanoid(),
            kb_id=kb_id,
            filename=file.filename,
            content=text,
            file_type=file_ext if file_ext else ".txt",
            file_size=file_size,
            schema_json=effective_schema,
            has_graph=False,
        )
        db.add(file_record)
        await db.flush()
        await db.refresh(file_record)

        results.append(
            FileUploadResponse(
                id=file_record.id,
                filename=file.filename,
                file_type=file_ext if file_ext else ".txt",
                file_size=file_size,
                has_graph=False,
            )
        )

    return results


async def construct_file_graph_service(
    file_id: str, client_id: str, db: AsyncSession
):
    if not GRAPHRAG_AVAILABLE:
        raise Exception("GraphRAG components not available.")

    await send_progress_update(client_id, "construction", 2, "清理旧缓存...")
    file_record = await KnowledgeBaseFileService.get_by_id(db, file_id)
    if not file_record:
        raise Exception("File not found")

    await clear_cache_files(file_record.kb_id, file_id)
    await send_progress_update(client_id, "construction", 5, "准备语料...")

    cfg = get_config_instance()
    schema = _ensure_schema(file_record.schema_json)

    documents = [{"title": file_record.filename, "text": file_record.content}]

    await send_progress_update(client_id, "construction", 10, "加载配置...")
    builder = constructor.KTBuilder(
        file_id, None, mode=cfg.construction.mode, config=cfg,
        schema_data=schema,
    )

    await send_progress_update(client_id, "construction", 20, "开始实体关系抽取...")

    def build_graph_sync():
        return builder.build_knowledge_graph(documents=documents)

    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, build_graph_sync)

    await send_progress_update(client_id, "construction", 85, "保存图谱数据...")
    chunks_data = dict(getattr(builder, 'all_chunks', {})) or None
    graph_data = builder.format_output() if hasattr(builder, 'format_output') else None

    from app.base_model import generate_nanoid

    existing_graph = await KnowledgeGraphService.get_by_file(db, file_id)
    if existing_graph:
        existing_graph.graph_data = graph_data
        existing_graph.chunks_data = chunks_data
        await db.flush()
    else:
        graph_record = KGraphModel(
            id=generate_nanoid(),
            file_id=file_id,
            graph_data=graph_data,
            chunks_data=chunks_data,
        )
        db.add(graph_record)
        await db.flush()
        await db.refresh(graph_record)

    await KnowledgeBaseFileService.update_file_graph_status(db, file_id, True)

    await db.commit()

    await send_progress_update(client_id, "construction", 95, "准备可视化数据...")
    graph_vis_data = await prepare_graph_visualization_from_data(graph_data)
    await send_progress_update(client_id, "construction", 100, "图谱构建完成!")

    try:
        await manager.send_message(
            {
                "type": "complete",
                "stage": "construction",
                "message": "图谱构建完成!",
                "timestamp": datetime.now().isoformat(),
            },
            client_id,
        )
    except Exception as e:
        logger.warning(f"Failed to send completion message: {e}")

    return graph_vis_data


# ──────────────────────────────────────────────
# Visualization helpers (kept from original)
# ──────────────────────────────────────────────
def prepare_subquery_visualization(sub_questions: List[Dict], reasoning_steps: List[Dict]) -> Dict:
    subquery_nodes = []
    subquery_links = []
    for idx, sq in enumerate(sub_questions):
        q_text = sq.get("sub-question", f"子问题 {idx + 1}")
        nid = f"sq_{idx}"
        subquery_nodes.append({"id": nid, "name": q_text[:30], "category": "sub_question", "symbolSize": 20})
        step_data = reasoning_steps[idx] if idx < len(reasoning_steps) else {}
        triples = step_data.get("triples", [])[:5]
        for t_idx, triple in enumerate(triples):
            tid = f"triple_{idx}_{t_idx}"
            parts = [p.strip().strip("'[]") for p in triple.split(",")] if "," in triple else [triple]
            if len(parts) >= 3:
                subquery_nodes.append({"id": tid, "name": parts[1][:20], "category": "triple", "symbolSize": 10})
                subquery_links.append({"source": nid, "target": tid, "name": parts[0][:15], "value": 1})
    return {"nodes": subquery_nodes, "links": subquery_links}


def prepare_retrieved_graph_visualization(triples: List[str]) -> Dict:
    nodes = {}
    links = []
    for triple in triples:
        parts = [p.strip().strip("'[]") for p in triple.split(",")] if "," in triple else [triple, "", ""]
        if len(parts) >= 3:
            h, r, t = parts[0], parts[1], parts[2]
            if h and h not in nodes:
                nodes[h] = {"id": h, "name": h[:20], "category": "entity", "symbolSize": 15}
            if t and t not in nodes:
                nodes[t] = {"id": t, "name": t[:20], "category": "entity", "symbolSize": 15}
            links.append({"source": h, "target": t, "name": r[:15], "value": 1})
    return {"nodes": list(nodes.values()), "links": links, "categories": [{"name": "entity"}], "stats": {"total_nodes": len(nodes), "total_edges": len(links)}}


def prepare_reasoning_flow_visualization(reasoning_steps: List[Dict]) -> Dict:
    nodes = [{"id": "question", "name": "问题", "category": "question", "symbolSize": 25}]
    links = []
    for i, step in enumerate(reasoning_steps):
        sid = f"step_{i}"
        label = step.get("question", f"步骤 {i + 1}")[:20]
        nodes.append({"id": sid, "name": label, "category": step.get("type", "step"), "symbolSize": 20})
        links.append({"source": "question" if i == 0 else f"step_{i - 1}", "target": sid, "name": "→", "value": 1})
    return {"nodes": nodes, "links": links, "categories": [{"name": "question"}, {"name": "sub_question"}, {"name": "ircot_step"}]}


# ──────────────────────────────────────────────
# Ask question (streaming) — per file
# ──────────────────────────────────────────────
def _sse(**kwargs):
    return f"data: {json.dumps(kwargs, ensure_ascii=False)}\n\n"


async def ask_file_question_stream(file_id: str, question: str, client_id: str, db: Optional[AsyncSession] = None):
    if not GRAPHRAG_AVAILABLE:
        yield _sse(type="error", message="GraphRAG components not available.")
        return

    yield _sse(type="status", progress=10, message="加载图谱数据...")

    graph_record = None
    if db:
        graph_record = await KnowledgeGraphService.get_by_file(db, file_id)
    if not graph_record:
        yield _sse(type="error", message="Graph not found. Please construct graph first.")
        return
    graph_data_source = graph_record.graph_data
    if not graph_data_source:
        yield _sse(type="error", message="Graph data is empty.")
        return

    yield _sse(type="status", progress=20, message="初始化检索系统...")
    dataset_name = file_id
    file_record = await KnowledgeBaseFileService.get_by_id(db, file_id) if db else None
    schema = _ensure_schema(file_record.schema_json if file_record else None)

    yield _sse(type="status", progress=30, message="初始化检索系统...")

    cfg = get_config_instance()
    graphq = decomposer.GraphQ(dataset_name, config=None, schema_data=schema)
    chunks_data = None
    if db and graph_record and hasattr(graph_record, 'chunks_data'):
        chunks_data = graph_record.chunks_data
    kt_retriever = retriever.KTRetriever(
        dataset_name,
        '',
        recall_paths=cfg.retrieval.recall_paths,
        schema_path='',
        top_k=cfg.retrieval.top_k_filter,
        mode="agent",
        config=cfg,
        chunks_data=chunks_data,
        graph_data=graph_data_source,
    )

    yield _sse(type="status", progress=40, message="构建索引...")
    kt_retriever.build_indices()

    def _dedup(items):
        return list({x: None for x in items}.keys())

    def _merge_chunk_contents(ids, mapping):
        return [mapping.get(i, f"[Missing content for chunk {i}]") for i in ids]

    yield _sse(type="status", progress=50, message="问题分解...")
    try:
        decomposition = graphq.decompose(question, '')
        sub_questions = decomposition.get("sub_questions", [])
        involved_types = decomposition.get("involved_types", {})
    except Exception as e:
        logger.error(f"Decompose failed: {e}")
        sub_questions = [{"sub-question": question}]
        involved_types = {"nodes": [], "relations": [], "attributes": []}

    reasoning_steps = []
    all_triples = set()
    all_chunk_ids = set()
    all_chunk_contents: Dict[str, str] = {}

    yield _sse(type="status", progress=65, message="初始检索...")
    for _, sq in enumerate(sub_questions):
        sq_text = sq.get("sub-question", question)
        retrieval_results, elapsed = kt_retriever.process_retrieval_results(
            sq_text, top_k=cfg.retrieval.top_k_filter, involved_types=involved_types
        )
        triples = retrieval_results.get("triples", []) or []
        chunk_ids = retrieval_results.get("chunk_ids", []) or []
        chunk_contents = retrieval_results.get("chunk_contents", []) or []
        if isinstance(chunk_contents, dict):
            for cid, ctext in chunk_contents.items():
                all_chunk_contents[cid] = ctext
        else:
            for i_c, cid in enumerate(chunk_ids):
                if i_c < len(chunk_contents):
                    all_chunk_contents[cid] = chunk_contents[i_c]
        all_triples.update(triples)
        all_chunk_ids.update(chunk_ids)
        reasoning_steps.append({
            "type": "sub_question", "question": sq_text,
            "triples": triples[:10], "triples_count": len(triples),
            "chunks_count": len(chunk_ids), "processing_time": elapsed,
            "chunk_contents": list(all_chunk_contents.values())[:3],
        })

    yield _sse(type="metadata", sub_questions=sub_questions,
               triples=list(all_triples)[:20], chunks=list(all_chunk_contents.values())[:10])

    initial_triples = _dedup(list(all_triples))
    initial_chunk_ids = list(set(all_chunk_ids))
    initial_chunk_contents = _merge_chunk_contents(initial_chunk_ids, all_chunk_contents)
    context_initial = (
        "=== Triples ===\n" +
        "\n".join(initial_triples[:20])
        + "\n=== Chunks ===\n" +
        "\n".join(initial_chunk_contents[:10])
    )
    init_prompt = kt_retriever.generate_prompt(question=question, sub_question=sub_questions, context=context_initial)

    if not cfg.retrieval.agent.enable_ircot:
        yield _sse(type="answer_start")
        answer_tokens = []
        try:
            async for token in kt_retriever.generate_answer_stream(init_prompt):
                answer_tokens.append(token)
                yield _sse(type="token", phase="answer", text=token)
        except Exception as e:
            err = f"答案生成失败: {e}"
            answer_tokens.append(err)
            yield _sse(type="token", phase="answer", text=err)
        yield _sse(type="answer_end")
        final_triples = initial_triples[:20]
        think = {"type": "init", "question": question, "triples": initial_triples,
                 "triples_count": len(initial_triples), "chunks_count": len(initial_chunk_ids),
                 "processing_time": 0, "chunk_contents": initial_chunk_contents[:3],
                 "thought": "".join(answer_tokens)}
        reasoning_steps.append(think)
        visualization_data = {
            "subqueries": prepare_subquery_visualization(sub_questions, reasoning_steps),
            "knowledge_graph": prepare_retrieved_graph_visualization(final_triples),
            "reasoning_flow": prepare_reasoning_flow_visualization(reasoning_steps),
            "retrieval_details": {"total_triples": len(final_triples),
                                  "total_chunks": len(initial_chunk_contents),
                                  "sub_questions_count": len(sub_questions),
                                  "triples_by_subquery": [r.get("triples_count", 0) for r in reasoning_steps if r.get("type") == "sub_question"]},
        }
        yield _sse(type="reasoning_steps", data={"reasoning_steps": reasoning_steps})
        yield _sse(type="visualization", data=visualization_data)
        yield _sse(type="done", answer="".join(answer_tokens))
        return

    # IRCoT path
    yield _sse(type="reasoning_start", step=0)
    initial_answer_tokens = []
    try:
        async for token in kt_retriever.generate_answer_stream(init_prompt):
            initial_answer_tokens.append(token)
            yield _sse(type="token", phase="reasoning", text=token)
    except Exception as e:
        err = f"初始答案生成失败: {e}"
        initial_answer_tokens.append(err)
        yield _sse(type="token", phase="reasoning", text=err)
    initial_answer = "".join(initial_answer_tokens)
    reasoning_steps.append({"type": "init", "question": question, "triples": initial_triples,
                            "triples_count": len(initial_triples), "chunks_count": len(initial_chunk_ids),
                            "processing_time": 0, "chunk_contents": initial_chunk_contents[:3],
                            "thought": initial_answer})
    yield _sse(type="reasoning_steps", data={"reasoning_steps": reasoning_steps})
    yield _sse(type="reasoning_end", step=0, thought=initial_answer[:300], query=question,
               triples=list(all_triples)[:5], triples_count=len(all_triples),
               chunks_count=len(all_chunk_contents))

    thoughts = [f"Initial: {initial_answer[:200]}"]
    current_query = question
    yield _sse(type="ircot_start")
    max_steps = getattr(getattr(cfg.retrieval, "agent", object()), "max_steps", 3)
    final_answer = None

    for step in range(1, max_steps + 1):
        loop_triples = _dedup(list(all_triples))
        loop_chunk_ids = list(set(all_chunk_ids))
        loop_chunk_contents = _merge_chunk_contents(loop_chunk_ids, all_chunk_contents)
        loop_ctx = "=== Triples ===\n" + "\n".join(loop_triples[:20]) + "\n=== Chunks ===\n" + "\n".join(loop_chunk_contents[:10])
        loop_prompt = f"""
        You are an expert knowledge assistant using iterative retrieval with chain-of-thought reasoning.
        Current Question: {question}
        Current Iteration Query: {current_query}
        Knowledge Context:\n{loop_ctx}
        Previous Thoughts: {" | ".join(thoughts) if thoughts else "None"}
        Instructions:
        1. If enough info answer with: So the answer is: <answer>
        2. Else propose new query with: The new query is: <query>
        Your reasoning:
        """
        yield _sse(type="reasoning_start", step=step)
        reasoning_tokens = []
        try:
            async for token in kt_retriever.generate_answer_stream(loop_prompt):
                reasoning_tokens.append(token)
                yield _sse(type="token", phase="reasoning", text=token)
        except Exception as e:
            err = f"推理错误: {e}"
            reasoning_tokens.append(err)
            yield _sse(type="token", phase="reasoning", text=err)
        reasoning = "".join(reasoning_tokens)
        thoughts.append(reasoning[:400])
        reasoning_steps.append({"type": "ircot_step", "question": current_query,
                                "triples": loop_triples[:10], "triples_count": len(loop_triples),
                                "chunks_count": len(loop_chunk_ids), "processing_time": 0,
                                "chunk_contents": loop_chunk_contents[:3], "thought": reasoning[:300]})
        yield _sse(type="reasoning_end", step=step, thought=reasoning[:300], query=current_query,
                   triples=loop_triples[:5], triples_count=len(loop_triples),
                   chunks_count=len(loop_chunk_ids))
        if "So the answer is:" in reasoning:
            import re
            m = re.search(r"So the answer is:\s*(.*)", reasoning, flags=re.IGNORECASE | re.DOTALL)
            final_answer = m.group(1).strip() if m else reasoning
            yield _sse(type="answer_found", answer=final_answer)
            break
        if "The new query is:" not in reasoning:
            final_answer = initial_answer or reasoning
            break
        new_query = reasoning.split("The new query is:", 1)[1].strip().splitlines()[0]
        if not new_query or new_query == current_query:
            final_answer = initial_answer or reasoning
            break
        current_query = new_query
        yield _sse(type="status", progress=min(90, 75 + step * 5), message=f"迭代检索 Step {step}...")
        try:
            new_ret, _ = kt_retriever.process_retrieval_results(current_query, top_k=cfg.retrieval.top_k_filter)
            new_triples = new_ret.get("triples", []) or []
            new_chunk_ids = new_ret.get("chunk_ids", []) or []
            new_cc = new_ret.get("chunk_contents", []) or []
            if isinstance(new_cc, dict):
                for cid, ct in new_cc.items():
                    all_chunk_contents[cid] = ct
            else:
                for i_c, cid in enumerate(new_chunk_ids):
                    if i_c < len(new_cc):
                        all_chunk_contents[cid] = new_cc[i_c]
            all_triples.update(new_triples)
            all_chunk_ids.update(new_chunk_ids)
        except Exception as e:
            logger.error(f"Iterative retrieval failed: {e}")
            break

    yield _sse(type="ircot_end")
    if final_answer is None:
        final_answer = initial_answer
    final_triples = _dedup(list(all_triples))[:20]
    final_chunk_ids = list(set(all_chunk_ids))
    final_chunk_contents = _merge_chunk_contents(final_chunk_ids, all_chunk_contents)[:10]
    visualization_data = {
        "subqueries": prepare_subquery_visualization(sub_questions, reasoning_steps),
        "knowledge_graph": prepare_retrieved_graph_visualization(final_triples),
        "reasoning_flow": prepare_reasoning_flow_visualization(reasoning_steps),
        "retrieval_details": {"total_triples": len(final_triples), "total_chunks": len(final_chunk_contents),
                              "sub_questions_count": len(sub_questions),
                              "triples_by_subquery": [r.get("triples_count", 0) for r in reasoning_steps if r.get("type") == "sub_question"]},
    }
    yield _sse(type="reasoning_steps", data={"reasoning_steps": reasoning_steps})
    yield _sse(type="visualization", data=visualization_data)
    yield _sse(type="done", answer=final_answer)


async def get_file_graph_service(file_id: str, db: AsyncSession):
    graph_record = await KnowledgeGraphService.get_by_file(db, file_id)
    if not graph_record or not graph_record.graph_data:
        return {"nodes": [], "links": [], "categories": [], "stats": {}}
    return await prepare_graph_visualization_from_data(graph_record.graph_data)


async def prepare_graph_visualization_from_data(graph_data) -> Dict:
    return _convert_visualization(graph_data)


def _convert_visualization(graph_data) -> Dict:
    if isinstance(graph_data, list):
        return convert_graphrag_format(graph_data)
    elif isinstance(graph_data, dict) and "nodes" in graph_data:
        return convert_standard_format(graph_data)
    return {"nodes": [], "links": [], "categories": [], "stats": {}}


def convert_graphrag_format(graph_data: list) -> Dict:
    nodes_dict = {}
    links = []
    for item in graph_data:
        if not isinstance(item, dict):
            continue
        start_node = item.get("start_node", {})
        end_node = item.get("end_node", {})
        relation = item.get("relation", "related_to")
        if start_node:
            sid = start_node.get("properties", {}).get("name", "")
            if sid and sid not in nodes_dict:
                nodes_dict[sid] = {
                    "id": sid, "name": sid[:30],
                    "category": start_node.get("properties", {}).get("schema_type", start_node.get("label", "entity")),
                    "symbolSize": 25, "properties": start_node.get("properties", {}),
                }
        if end_node:
            eid = end_node.get("properties", {}).get("name", "")
            if isinstance(eid, (list, dict)):
                eid = str(eid.get("name", "")) if isinstance(eid, dict) else str(eid)
            if eid and eid not in nodes_dict:
                nodes_dict[eid] = {
                    "id": eid, "name": eid[:30],
                    "category": end_node.get("properties", {}).get("schema_type", end_node.get("label", "entity")),
                    "symbolSize": 25, "properties": end_node.get("properties", {}),
                }
        if sid and eid:
            links.append({"source": sid, "target": eid, "name": relation, "value": 1})
    categories = list(set(n["category"] for n in nodes_dict.values() if n["category"]))
    return {"nodes": list(nodes_dict.values()), "links": links,
            "categories": [{"name": c} for c in categories],
            "stats": {"total_nodes": len(nodes_dict), "total_edges": len(links),
                      "displayed_nodes": len(nodes_dict), "displayed_edges": len(links)}}


def convert_standard_format(graph_data: Dict) -> Dict:
    nodes = graph_data.get("nodes", [])
    edges = graph_data.get("edges", []) or graph_data.get("links", [])
    categories = graph_data.get("categories", [])
    for n in nodes:
        if "symbolSize" not in n:
            n["symbolSize"] = 25
    for e in edges:
        if "value" not in e:
            e["value"] = 1
    return {"nodes": nodes, "links": edges, "categories": categories,
            "stats": {"total_nodes": len(nodes), "total_edges": len(edges),
                      "displayed_nodes": len(nodes), "displayed_edges": len(edges)}}
