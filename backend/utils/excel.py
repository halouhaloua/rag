#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Author: 臧成龙
@Contact: 939589097@qq.com
@Time: 2025-12-31
@File: excel.py
@Desc: Excel处理工具类 - 
"""
from io import BytesIO
from typing import List, Dict, Any, Type, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pydantic import BaseModel


class ExcelHandler:
    """Excel处理工具类"""
    
    # 表头样式
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    
    # 边框样式
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    @classmethod
    def export_to_excel(
        cls,
        data: List[Dict[str, Any]],
        columns: Dict[str, str],
        sheet_name: str = "Sheet1"
    ) -> BytesIO:
        """
        导出数据到Excel
        :param data: 数据列表，每个元素是一个字典
        :param columns: 列映射，格式为 {字段名: 显示名}
        :param sheet_name: 工作表名称
        :return: Excel文件的BytesIO对象
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 写入表头
        headers = list(columns.values())
        field_names = list(columns.keys())
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.HEADER_ALIGNMENT
            cell.border = cls.THIN_BORDER
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, field in enumerate(field_names, 1):
                value = row_data.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = cls.THIN_BORDER
                cell.alignment = Alignment(vertical="center")
        
        # 自动调整列宽
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @classmethod
    def import_from_excel(
        cls,
        file_content: bytes,
        columns: Dict[str, str],
        schema: Optional[Type[BaseModel]] = None
    ) -> List[Dict[str, Any]]:
        """
        从Excel导入数据
        :param file_content: Excel文件内容
        :param columns: 列映射，格式为 {字段名: 显示名}
        :param schema: 可选的Pydantic Schema用于数据验证
        :return: 数据列表
        """
        wb = load_workbook(filename=BytesIO(file_content), read_only=True)
        ws = wb.active
        
        # 读取表头，建立显示名到字段名的映射
        header_to_field = {v: k for k, v in columns.items()}
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        
        # 第一行是表头
        headers = rows[0]
        field_indices = {}
        for idx, header in enumerate(headers):
            if header in header_to_field:
                field_indices[idx] = header_to_field[header]
        
        # 读取数据行
        result = []
        for row in rows[1:]:
            if not any(row):  # 跳过空行
                continue
            
            row_data = {}
            for idx, field_name in field_indices.items():
                value = row[idx] if idx < len(row) else None
                row_data[field_name] = value
            
            # 如果提供了schema，进行数据验证
            if schema:
                try:
                    validated = schema(**row_data)
                    row_data = validated.model_dump()
                except Exception:
                    continue  # 跳过验证失败的行
            
            result.append(row_data)
        
        wb.close()
        return result
    
    @classmethod
    def generate_template(
        cls,
        columns: Dict[str, str],
        sheet_name: str = "Sheet1"
    ) -> BytesIO:
        """
        生成导入模板
        :param columns: 列映射，格式为 {字段名: 显示名}
        :param sheet_name: 工作表名称
        :return: Excel模板文件的BytesIO对象
        """
        return cls.export_to_excel([], columns, sheet_name)
